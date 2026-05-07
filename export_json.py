"""
export_json.py
讀取 trading-journal (2) - 每日損益重算.xlsx，計算所有績效指標，
輸出 docs/data.json 供 GitHub Pages 網站使用。

使用流程：
  1. python rebuild_daily_pnl.py    → 更新 XLSX
  2. python export_json.py          → 輸出 docs/data.json
  3. git add docs/data.json && git commit -m "update" && git push
"""
from __future__ import annotations

import json
import math
import statistics
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

INPUT_FILE = Path("trading-journal (2) - 每日損益重算.xlsx")
OUTPUT_JSON = Path("docs/data.json")
CAPITAL = 100_000.0
RF_ANNUAL = 0.017          # 無風險利率（年）
RF_DAILY  = RF_ANNUAL / 252


# ── helpers ──────────────────────────────────────────────────────────────────

def as_date(v) -> Optional[date]:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def as_float(v, default: float = 0.0) -> float:
    if v in (None, ""):
        return default
    return float(v)


def rnd(v, d=2) -> float:
    return round(float(v), d)


# ── readers ──────────────────────────────────────────────────────────────────

def read_equity(wb) -> list[dict]:
    """每日損益重算 工作表，第 6 列起為資料。"""
    ws = wb["每日損益重算"]
    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        dt = as_date(row[0])
        if dt is None:
            continue
        rows.append({
            "date":           dt.isoformat(),
            "cum_realized":   rnd(as_float(row[2])),
            "unrealized":     rnd(as_float(row[3])),
            "total_pnl":      rnd(as_float(row[4])),
            "cash":           rnd(as_float(row[5])),
            "position_value": rnd(as_float(row[6])),
            "equity":         rnd(as_float(row[7])),
            "daily_change":   rnd(as_float(row[8])),   # 第一列 None → 0.0
            "num_positions":  int(as_float(row[9])),
        })
    return rows


def read_trades(wb) -> list[dict]:
    """交易記錄 工作表，第 3 列起為資料。"""
    ws = wb["交易記錄"]
    trades = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = row[1]
        if code is None:
            continue
        # 股票代號轉字串
        if isinstance(code, float):
            code = int(code)
        ticker = str(code)

        # 名稱去除標記符號（如 * # 等）
        raw_name = str(row[2] or "").strip()
        name = raw_name.rstrip("*＊#")

        buy_date  = as_date(row[5])
        sell_date = as_date(row[6])
        days = (sell_date - buy_date).days if sell_date and buy_date else 0

        pnl = row[13]   # 已實現損益（None = 尚未實現）
        ret = row[14]   # 報酬率%（decimal，如 0.1177）

        trades.append({
            "ticker":     ticker,
            "name":       name,
            "market":     str(row[3] or ""),
            "industry":   str(row[4] or ""),
            "buy_date":   buy_date.isoformat()  if buy_date  else "",
            "sell_date":  sell_date.isoformat() if sell_date else "",
            "status":     str(row[7] or ""),
            "qty":        as_float(row[8]),
            "buy_price":  rnd(as_float(row[9]),  4),
            "sell_price": rnd(as_float(row[10]), 4) if row[10] not in (None, "") else 0.0,
            "buy_fee":    as_float(row[11]),
            "sell_fee":   as_float(row[12]),
            "pnl":        rnd(as_float(pnl)) if pnl not in (None, "") else 0.0,
            "ret":        rnd(as_float(ret), 4) if ret not in (None, "") else 0.0,
            "days":       days,
            "note":       str(row[16] or ""),
        })
    return trades


# ── metrics ──────────────────────────────────────────────────────────────────

def calc_metrics(equity_rows: list[dict], trades: list[dict]) -> dict:
    equities = [r["equity"] for r in equity_rows]
    if not equities:
        raise ValueError("equity 資料為空")

    # 日報酬率（從第 2 日開始）
    returns = []
    for i in range(1, len(equity_rows)):
        prev = equities[i - 1]
        if prev > 0:
            returns.append((equities[i] - prev) / prev)

    trading_days = len(equity_rows)
    final_equity = equities[-1]
    total_pnl    = equity_rows[-1]["total_pnl"]
    total_ret    = total_pnl / CAPITAL

    # 年化報酬率
    ann_ret = (1 + total_ret) ** (252 / trading_days) - 1 if trading_days > 0 else 0.0

    # 年化波動度 / Sharpe / Sortino
    if len(returns) >= 2:
        std_daily = statistics.stdev(returns)           # sample std
        ann_vol   = std_daily * math.sqrt(252)

        # Sharpe = (年化報酬 - 無風險利率) / 年化波動度
        sharpe = (ann_ret - RF_ANNUAL) / ann_vol if ann_vol > 0 else 0.0

        # Sortino：分母只計下行偏差（負超額報酬的 sample std）
        down_rets = [r for r in returns if r < RF_DAILY]
        if len(down_rets) >= 2:
            down_std_daily = statistics.stdev(down_rets)
            sortino = (ann_ret - RF_ANNUAL) / (down_std_daily * math.sqrt(252)) if down_std_daily > 0 else 0.0
        else:
            sortino = sharpe
    else:
        ann_vol = sharpe = sortino = 0.0

    # 最大回撤（MDD）
    peak = equities[0]
    mdd  = 0.0
    for eq in equities:
        if eq > peak:
            peak = eq
        dd = (eq - peak) / peak
        if dd < mdd:
            mdd = dd

    # 交易統計（僅計算已平倉且有 pnl 的筆數）
    closed = [t for t in trades if t["sell_date"] and t["pnl"] != 0.0]
    wins   = [t for t in closed if t["pnl"] > 0]
    losses = [t for t in closed if t["pnl"] < 0]

    total_trades = len(closed)
    win_rate  = len(wins) / total_trades if total_trades > 0 else 0.0
    avg_win   = sum(t["pnl"] for t in wins)   / len(wins)   if wins   else 0.0
    avg_loss  = sum(t["pnl"] for t in losses) / len(losses) if losses else 0.0

    # Profit factor = avg_win / |avg_loss|（與儀表板口徑一致）
    profit_factor   = avg_win / abs(avg_loss) if avg_loss != 0 else 0.0
    expected_value  = win_rate * avg_win + (1 - win_rate) * avg_loss

    return {
        "capital":        CAPITAL,
        "final_equity":   rnd(final_equity),
        "total_pnl":      rnd(total_pnl),
        "total_ret":      rnd(total_ret,  4),
        "ann_ret":        rnd(ann_ret,    4),
        "ann_vol":        rnd(ann_vol,    4),
        "sharpe":         rnd(sharpe,     2),
        "sortino":        rnd(sortino,    2),
        "mdd":            rnd(mdd,        4),
        "trading_days":   trading_days,
        "total_trades":   total_trades,
        "win_trades":     len(wins),
        "loss_trades":    len(losses),
        "win_rate":       rnd(win_rate,   4),
        "avg_win":        rnd(avg_win,    2),
        "avg_loss":       rnd(avg_loss,   2),
        "profit_factor":  rnd(profit_factor, 2),
        "expected_value": rnd(expected_value, 2),
    }


# ── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"找不到 {INPUT_FILE}，請先執行 python rebuild_daily_pnl.py"
        )

    print(f"讀取 {INPUT_FILE} …")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)

    equity_rows = read_equity(wb)
    trades      = read_trades(wb)
    wb.close()

    print(f"  equity 資料：{len(equity_rows)} 列")
    print(f"  交易記錄：{len(trades)} 筆（含持倉）")

    metrics = calc_metrics(equity_rows, trades)

    data = {
        "equity":  equity_rows,
        "trades":  trades,
        "capital": CAPITAL,
        "metrics": metrics,
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    print(f"\n[OK] 匯出完成 -> {OUTPUT_JSON}")
    print(f"  最終權益：NT$ {metrics['final_equity']:,.2f}")
    print(f"  總報酬率：{metrics['total_ret']*100:.2f}%")
    print(f"  夏普值：  {metrics['sharpe']}")
    print(f"  最大回撤：{metrics['mdd']*100:.2f}%")


if __name__ == "__main__":
    main()
