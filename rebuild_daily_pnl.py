from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE = Path("trading-journal (2).xlsx")
OUTPUT_FILE = Path("trading-journal (2) - 每日損益重算.xlsx")
INITIAL_CAPITAL = 100000.0


# 補上活頁簿內缺失的興櫃收盤價。
# 來源：
# - 3585 聯致：Goodinfo 2026/04/15 歷史股價頁面
# - 7828 創新服務：Goodinfo 2026/04/15 歷史股價頁面
# 其中 7828 在 2026/04/16 找不到同日收盤，以下一交易日前最後可得收盤價 1325 承接。
SUPPLEMENTAL_CLOSES: Dict[int, Dict[date, float]] = {
    3585: {
        date(2026, 4, 13): 30.1,
        date(2026, 4, 14): 49.2,
        date(2026, 4, 15): 62.0,
    },
    7828: {
        date(2026, 4, 1): 1310.0,
        date(2026, 4, 2): 1285.0,
        date(2026, 4, 7): 1325.0,
        date(2026, 4, 8): 1405.0,
        date(2026, 4, 9): 1415.0,
        date(2026, 4, 10): 1395.0,
        date(2026, 4, 13): 1325.0,
        date(2026, 4, 14): 1315.0,
        date(2026, 4, 15): 1325.0,
    },
}


@dataclass
class Trade:
    trade_id: int
    code: int
    name: str
    market: str
    industry: str
    buy_date: date
    sell_date: Optional[date]
    shares: float
    buy_px: float
    sell_px: Optional[float]
    buy_fee: float
    sell_fee_tax: float
    realized: Optional[float]
    note: Optional[str]

    @property
    def buy_cost(self) -> float:
        return self.shares * self.buy_px + self.buy_fee

    @property
    def net_sell_proceeds(self) -> Optional[float]:
        if self.realized is None:
            return None
        return self.buy_cost + self.realized


def as_date(value) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        raw = str(int(value))
        if len(raw) == 8:
            return date(int(raw[:4]), int(raw[4:6]), int(raw[6:8]))
    raise ValueError(f"Unsupported date value: {value!r}")


def as_code(value) -> Optional[int]:
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, int):
        return value
    return int(str(value).strip())


def as_float(value, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    return float(value)


def load_trades(workbook) -> List[Trade]:
    ws = workbook["交易記錄"]
    trades: List[Trade] = []
    for row in range(3, ws.max_row + 1):
        code = as_code(ws.cell(row, 2).value)
        if code is None:
            continue
        trade = Trade(
            trade_id=int(as_float(ws.cell(row, 1).value)),
            code=code,
            name=str(ws.cell(row, 3).value or ""),
            market=str(ws.cell(row, 4).value or ""),
            industry=str(ws.cell(row, 5).value or ""),
            buy_date=as_date(ws.cell(row, 6).value),
            sell_date=as_date(ws.cell(row, 7).value),
            shares=as_float(ws.cell(row, 9).value),
            buy_px=as_float(ws.cell(row, 10).value),
            sell_px=(
                None
                if ws.cell(row, 11).value in (None, "")
                else float(ws.cell(row, 11).value)
            ),
            buy_fee=as_float(ws.cell(row, 12).value),
            sell_fee_tax=as_float(ws.cell(row, 13).value),
            realized=(
                None
                if ws.cell(row, 14).value in (None, "")
                else float(ws.cell(row, 14).value)
            ),
            note=ws.cell(row, 17).value,
        )
        trades.append(trade)
    return trades


def load_local_closes(workbook) -> Tuple[Dict[int, Dict[date, float]], List[date]]:
    ws = workbook["收盤價"]
    closes: Dict[int, Dict[date, float]] = {}
    trading_dates = set()
    for row in range(2, ws.max_row + 1):
        dt = as_date(ws.cell(row, 1).value)
        code = as_code(ws.cell(row, 2).value)
        px = ws.cell(row, 3).value
        if dt is None or code is None or px in (None, ""):
            continue
        trading_dates.add(dt)
        closes.setdefault(code, {})[dt] = float(px)
    return closes, sorted(trading_dates)


def merge_supplemental_closes(
    closes: Dict[int, Dict[date, float]]
) -> Dict[int, Dict[date, float]]:
    merged = {code: series.copy() for code, series in closes.items()}
    for code, series in SUPPLEMENTAL_CLOSES.items():
        merged.setdefault(code, {}).update(series)
    return merged


def price_as_of(
    close_series: Dict[date, float], current_date: date
) -> Tuple[Optional[float], Optional[date]]:
    eligible = [dt for dt in close_series if dt <= current_date]
    if not eligible:
        return None, None
    latest = max(eligible)
    return close_series[latest], latest


def style_header(ws, header_row: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autofit_columns(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 24)


def remove_sheet_if_exists(workbook, name: str) -> None:
    if name in workbook.sheetnames:
        del workbook[name]


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(INPUT_FILE)

    wb_values = load_workbook(INPUT_FILE, data_only=True)
    trades = load_trades(wb_values)
    local_closes, local_dates = load_local_closes(wb_values)
    closes = merge_supplemental_closes(local_closes)

    start_date = date(2025, 12, 24)
    end_date = max(
        max(local_dates),
        max(trade.buy_date for trade in trades),
        max(
            trade.sell_date
            for trade in trades
            if trade.sell_date is not None
        ),
    )
    event_dates = {
        trade.buy_date
        for trade in trades
        if start_date <= trade.buy_date <= end_date
    } | {
        trade.sell_date
        for trade in trades
        if trade.sell_date is not None and start_date <= trade.sell_date <= end_date
    }
    calendar = sorted(
        {
            dt
            for dt in local_dates
            if start_date <= dt <= end_date
        }
        | event_dates
    )

    missing_codes = sorted(
        {
            trade.code
            for trade in trades
            if trade.code not in closes
        }
    )
    if missing_codes:
        raise ValueError(f"Missing close series for codes: {missing_codes}")

    wb = load_workbook(INPUT_FILE)
    remove_sheet_if_exists(wb, "每日損益重算")
    remove_sheet_if_exists(wb, "每日明細重算")

    summary_ws = wb.create_sheet("每日損益重算")
    detail_ws = wb.create_sheet("每日明細重算")

    summary_ws["A1"] = "每日損益重算"
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws["A2"] = "計算口徑：期末權益 = 現金餘額 + 持倉市值；未實現損益含買進手續費。"
    summary_ws["A3"] = (
        "補價說明：3585、7828 依 Goodinfo 歷史股價補齊；7828 在 2026-04-16 "
        "缺同日收盤，沿用 2026-04-15 收盤 1325。"
    )

    summary_headers = [
        "日期",
        "當日已實現損益",
        "累積已實現損益",
        "期末未實現損益",
        "總損益",
        "現金餘額",
        "持倉市值",
        "權益總值",
        "日權益變動",
        "持倉檔數",
        "已收盤價覆蓋說明",
    ]
    summary_ws.append([])
    summary_ws.append(summary_headers)
    style_header(summary_ws, 5)

    detail_ws["A1"] = "每日明細重算"
    detail_ws["A1"].font = Font(bold=True, size=14)
    detail_ws["A2"] = "同日賣出者，當天列入已實現損益，期末未實現損益為 0。"
    detail_headers = [
        "日期",
        "交易ID",
        "股票代號",
        "股票名稱",
        "市場別",
        "產業別",
        "事件",
        "股數",
        "買進日期",
        "賣出日期",
        "買進均價",
        "當日收盤價",
        "收盤價日期",
        "當日持倉市值",
        "當日已實現損益",
        "期末未實現損益",
        "當日說明",
    ]
    detail_ws.append([])
    detail_ws.append(detail_headers)
    style_header(detail_ws, 4)

    cash = INITIAL_CAPITAL
    cumulative_realized = 0.0
    detail_rows: List[List] = []
    buy_events: Dict[date, List[Trade]] = {}
    sell_events: Dict[date, List[Trade]] = {}
    for trade in trades:
        buy_events.setdefault(trade.buy_date, []).append(trade)
        if trade.sell_date:
            sell_events.setdefault(trade.sell_date, []).append(trade)

    for trade in trades:
        if trade.buy_date < start_date:
            cash -= trade.buy_cost
        if trade.sell_date and trade.sell_date < start_date:
            proceeds = trade.net_sell_proceeds
            if proceeds is None:
                raise ValueError(f"Trade {trade.trade_id} missing realized pnl")
            cash += proceeds
            cumulative_realized += trade.realized or 0.0

    prev_equity: Optional[float] = None

    for current_date in calendar:
        for trade in buy_events.get(current_date, []):
            cash -= trade.buy_cost

        day_realized = 0.0
        for trade in sell_events.get(current_date, []):
            proceeds = trade.net_sell_proceeds
            if proceeds is None:
                raise ValueError(f"Trade {trade.trade_id} missing realized pnl")
            cash += proceeds
            day_realized += trade.realized or 0.0
            cumulative_realized += trade.realized or 0.0

        open_positions = [
            trade
            for trade in trades
            if trade.buy_date <= current_date
            and (trade.sell_date is None or current_date < trade.sell_date)
        ]

        position_value = 0.0
        unrealized_total = 0.0
        carry_notes: List[str] = []

        for trade in open_positions:
            series = closes[trade.code]
            close_px, close_dt = price_as_of(series, current_date)
            if close_px is None or close_dt is None:
                raise ValueError(
                    f"No price available for {trade.code} on or before {current_date.isoformat()}"
                )
            market_value = trade.shares * close_px
            unrealized = market_value - trade.buy_cost
            position_value += market_value
            unrealized_total += unrealized
            if close_dt != current_date:
                carry_notes.append(f"{trade.code}沿用{close_dt.isoformat()}收盤")
            detail_rows.append(
                [
                    current_date,
                    trade.trade_id,
                    trade.code,
                    trade.name,
                    trade.market,
                    trade.industry,
                    "持有中",
                    trade.shares,
                    trade.buy_date,
                    trade.sell_date,
                    trade.buy_px,
                    close_px,
                    close_dt,
                    market_value,
                    0.0,
                    unrealized,
                    "當日期末持有部位",
                ]
            )

        for trade in sell_events.get(current_date, []):
            detail_rows.append(
                [
                    current_date,
                    trade.trade_id,
                    trade.code,
                    trade.name,
                    trade.market,
                    trade.industry,
                    "當日賣出",
                    trade.shares,
                    trade.buy_date,
                    trade.sell_date,
                    trade.buy_px,
                    trade.sell_px,
                    current_date,
                    0.0,
                    trade.realized or 0.0,
                    0.0,
                    "已由未實現轉入已實現",
                ]
            )

        total_pnl = cumulative_realized + unrealized_total
        equity = cash + position_value
        day_change = None if prev_equity is None else equity - prev_equity
        diff = round(total_pnl - (equity - INITIAL_CAPITAL), 6)
        if abs(diff) > 0.01:
            raise ValueError(
                f"PnL mismatch on {current_date.isoformat()}: "
                f"pnl={total_pnl}, equity_delta={equity - INITIAL_CAPITAL}"
            )
        prev_equity = equity

        coverage_note = "；".join(sorted(set(carry_notes))) if carry_notes else ""
        summary_ws.append(
            [
                current_date,
                day_realized,
                cumulative_realized,
                unrealized_total,
                total_pnl,
                cash,
                position_value,
                equity,
                day_change,
                len(open_positions),
                coverage_note,
            ]
        )

    for row in detail_rows:
        detail_ws.append(row)

    date_fmt = "yyyy-mm-dd"
    money_fmt = '#,##0.00;(#,##0.00)'
    int_fmt = '#,##0'

    for row in summary_ws.iter_rows(min_row=6, max_row=summary_ws.max_row):
        row[0].number_format = date_fmt
        for cell in row[1:9]:
            cell.number_format = money_fmt
        row[9].number_format = int_fmt

    for row in detail_ws.iter_rows(min_row=5, max_row=detail_ws.max_row):
        row[0].number_format = date_fmt
        row[8].number_format = date_fmt
        row[9].number_format = date_fmt
        row[12].number_format = date_fmt
        row[7].number_format = '#,##0.####'
        for idx in [10, 11, 13, 14, 15]:
            row[idx].number_format = money_fmt

    summary_ws.freeze_panes = "A6"
    detail_ws.freeze_panes = "A5"
    autofit_columns(summary_ws)
    autofit_columns(detail_ws)

    wb.save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE}")
    print(f"Summary rows: {summary_ws.max_row - 5}")
    print(f"Detail rows: {detail_ws.max_row - 4}")


if __name__ == "__main__":
    main()
